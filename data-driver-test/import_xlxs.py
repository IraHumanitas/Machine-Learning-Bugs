import json
import openpyxl
from openpyxl import Workbook
import os

def download_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "test_case", "description", "username", "password", "button_name", "button_type", 
        "expected_selector", "expected_value", "url", "input_action_username", "input_action_password", 
        "actual_data_username", "actual_data_password"
    ]
    ws.append(headers)

    filename = "test_cases_format.xlsx"
    wb.save(filename)
    print(f"Format Excel berhasil diunduh: {filename}")

def import_xlsx_to_json():
    filename = input("Masukkan nama file XLSX untuk diimpor (misalnya 'test_cases_format.xlsx'): ")
    
    if not os.path.exists(filename):
        print("File tidak ditemukan!")
        return

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True): 
        test_case_data = {
            "test_case": row[0],
            "description": row[1],
            "selector_input": {
                "username": row[2],
                "password": row[3]
            },
            "button_selector": {
                "button_name": row[4],
                "button_type": row[5]
            },
            "expected_result": {
                "expected_selector": row[6],
                "expected_value": row[7],
                "url": row[8]
            },
            "input_action": {
                "username": row[9],
                "password": row[10]
            },
            "actual_data": {
                "username": row[11],
                "password": row[12]
            }
        }
        data.append(test_case_data)

    json_filename = "test_cases.json"
    if os.path.exists(json_filename):
        with open(json_filename, 'r') as f:
            existing_data = json.load(f)
    else:
        existing_data = []

    existing_data.extend(data)

    with open(json_filename, 'w') as f:
        json.dump(existing_data, f, indent=4)

    print(f"Data berhasil diimpor dan disimpan ke {json_filename}")

def main():
    print("Pilih opsi:")
    print("1. Download format XLSX")
    print("2. Import data dari XLSX ke JSON")
    choice = input("Masukkan pilihan (1 atau 2): ")

    if choice == "1":
        download_xlsx()
    elif choice == "2":
        import_xlsx_to_json()
    else:
        print("Pilihan tidak valid!")

if __name__ == "__main__":
    main()
